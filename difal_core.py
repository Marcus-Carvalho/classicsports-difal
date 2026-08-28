# -*- coding: utf-8 -*-
"""
Classic Sports - Guias DIFAL
Motor de extracao (sem interface grafica).

Le os ZIPs de guias baixados da Secretaria da Fazenda, identifica a linha
digitavel de cada PDF (GNRE ou DARE-SP) e gera a planilha para pagamento
em massa no banco.

Regra de ouro: o codigo exportado e a COPIA LITERAL da linha digitavel
impressa no PDF, apenas sem espacos e sem hifens. Nada e recalculado,
corrigido ou inventado.
"""

import os
import re
import sys
import json
import shutil
import zipfile
import tempfile
import unicodedata
from pathlib import Path
from datetime import datetime

# ---------------------------------------------------------------- empresas

# codigo do ZIP -> nome oficial da empresa
EMPRESAS = {
    "127": "ADVENTUREX",
    "128": "ADRENALINEX",
    "132": "J A TERRA",
    "133": "F F DE SOUZA",
    "199": "SOARES SOARES DE CARVALHO",
    "249": "DL MOTOS",
    "263": "CARVALHO SOARES DE CARVALHO",
    "273": "CLASSIC SPORTS EXPRESS",
    "296": "AM RACING",
    "305": "CLASSIC SPORTS EXPRESS 2",
    "308": "DM MOTOS",
    "309": "BL MOTOS",
    "311": "PLANET MOTORS",
    "317": "CLASSIC BARRACAO",
    "324": "AM RACING 2",
    "331": "M V S DE CARVALHO",
    "332": "J ANTUNES TERRA",
    "341": "G JOSE DE CARVALHO",
    "52": "R A S DE CARVALHO",
    "7": "M V SOARES",
    "73": "G J DE CARVALHO",
}

TOTAL_EMPRESAS = len(EMPRESAS)


def _norm(texto):
    """minusculas, sem acento, so letras e numeros"""
    t = unicodedata.normalize("NFD", str(texto).lower())
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9]", "", t)


_EMPRESAS_NORM = {_norm(v): (k, v) for k, v in EMPRESAS.items()}


def empresa_do_zip(nome_arquivo):
    """
    Descobre a empresa a partir do nome do ZIP.
    Ex.: notas_gnre_199_soares_soares_de_carvalho__2_.zip  ->  ('199', 'SOARES SOARES DE CARVALHO')
         128_adrenalinex(3).zip                            ->  ('128', 'ADRENALINEX')
    Retorna (codigo, nome, reconhecida:boolean).
    """
    base = Path(nome_arquivo).stem
    base = re.sub(r"\(\d+\)", "", base)          # ignora (1), (2), (3)...
    base = re.sub(r"__\d+_$", "", base)          # ignora sufixo __2_
    partes = [p for p in re.split(r"[_\-\s]+", base) if p]

    # 1) primeiro token puramente numerico = codigo da empresa
    for p in partes:
        if p.isdigit() and p in EMPRESAS:
            return p, EMPRESAS[p], True

    # 2) fallback: nome da empresa dentro do nome do arquivo
    alvo = _norm(base)
    melhor = None
    for chave, (cod, nome) in _EMPRESAS_NORM.items():
        if chave and chave in alvo:
            if melhor is None or len(chave) > len(melhor[0]):
                melhor = (chave, cod, nome)
    if melhor:
        return melhor[1], melhor[2], True

    # 3) nao reconhecida
    numeros = [p for p in partes if p.isdigit()]
    codigo = numeros[0] if numeros else ""
    return codigo, base.upper(), False


# ------------------------------------------------------- linha digitavel

# DARE-SP:  85810000000-5 07850185112-3 60590173537-9 62220260812-2
RE_DARE = re.compile(
    r"(?<!\d)(\d{11})-(\d)\s+(\d{11})-(\d)\s+(\d{11})-(\d)\s+(\d{11})-(\d)(?!\d)"
)
# GNRE:     85830000000 9 18000185112 0 60590083139 0 51420260430 8
RE_GNRE = re.compile(
    r"(?<!\d)(\d{11})\s+(\d)\s+(\d{11})\s+(\d)\s+(\d{11})\s+(\d)\s+(\d{11})\s+(\d)(?!\d)"
)
RE_VALOR = re.compile(r"R\$\s*([\d.]*\d,\d{2})")
RE_DATA = re.compile(r"(\d{2}/\d{2}/\d{4})")

ROTULO_VALOR = {"GNRE": "Total a Recolher", "DARE-SP": "08 - Valor Total"}
ROTULO_VENC = {"GNRE": "Documento Válido para pagamento até", "DARE-SP": "07 - Data de Vencimento"}


def _para_float(texto):
    return float(str(texto).replace(".", "").replace(",", "."))


def _texto_pdf(caminho, max_paginas=3):
    """Extrai texto do PDF ate achar a linha digitavel (normalmente pagina 1)."""
    import pdfplumber

    partes = []
    with pdfplumber.open(caminho) as pdf:
        for pagina in pdf.pages[:max_paginas]:
            try:
                partes.append(pagina.extract_text() or "")
            except Exception:
                partes.append("")
            texto = "\n".join(partes)
            if RE_DARE.search(texto) or RE_GNRE.search(texto):
                return texto
    return "\n".join(partes)


def processar_pdf(caminho):
    """
    Le um PDF e devolve dicionario com tipo, codigo (48 digitos), valor e vencimento.
    Nunca levanta excecao: erro vira campo 'motivo'.
    """
    caminho = Path(caminho)
    saida = {
        "arquivo": caminho.name,
        "tipo": None,
        "codigo": None,
        "valor": None,
        "vencimento": None,
        "valor_por_fallback": False,
        "motivo": None,
    }
    try:
        texto = _texto_pdf(caminho)
    except Exception as erro:
        saida["motivo"] = "falha ao ler o PDF: " + str(erro)
        return saida

    achado = RE_DARE.search(texto)
    if achado:
        saida["tipo"] = "DARE-SP"
    else:
        achado = RE_GNRE.search(texto)
        if achado:
            saida["tipo"] = "GNRE"

    if not achado:
        saida["motivo"] = "linha digitável não encontrada no PDF"
        return saida

    saida["codigo"] = "".join(achado.groups())   # 44 digitos + 4 DVs = 48

    if len(saida["codigo"]) != 48 or not saida["codigo"].isdigit():
        saida["motivo"] = "código com " + str(len(saida["codigo"])) + " dígitos (esperado 48)"
        return saida

    # valor: procura a partir do rotulo oficial do tipo de guia
    pos = texto.find(ROTULO_VALOR[saida["tipo"]])
    if pos >= 0:
        m = RE_VALOR.search(texto[pos:pos + 1500])
        if m:
            saida["valor"] = _para_float(m.group(1))
    if saida["valor"] is None:
        todos = RE_VALOR.findall(texto)
        if todos:
            saida["valor"] = _para_float(todos[-1])
            saida["valor_por_fallback"] = True
        else:
            saida["motivo"] = "valor não localizado"
            return saida

    # vencimento (apenas informativo, nao vai para a planilha)
    pos = texto.find(ROTULO_VENC[saida["tipo"]])
    if pos >= 0:
        m = RE_DATA.search(texto[pos:pos + 400])
        if m:
            saida["vencimento"] = m.group(1)

    return saida


# ------------------------------------------------------------ lote

class ResultadoLote(object):
    def __init__(self):
        self.guias = []            # dicts prontos para a planilha
        self.duplicadas = []       # mesmo codigo de barras dentro do lote
        self.sem_codigo = []       # PDFs sem linha digitavel
        self.nao_autorizadas = []  # ZIPs de empresa fora da lista
        self.zips = 0
        self.pdfs = 0
        self.inicio = datetime.now()

    @property
    def total(self):
        return sum(g["valor"] for g in self.guias)

    def por_empresa(self):
        agrupado = {}
        for g in self.guias:
            item = agrupado.setdefault(g["empresa"], {"guias": 0, "total": 0.0})
            item["guias"] += 1
            item["total"] += g["valor"]
        return dict(sorted(agrupado.items()))

    def por_tipo(self):
        agrupado = {}
        for g in self.guias:
            agrupado[g["tipo"]] = agrupado.get(g["tipo"], 0) + 1
        return agrupado


def extrair_zips(caminhos_zip, pasta_destino):
    """
    Extrai cada ZIP em uma subpasta limpa e devolve
    [(codigo, empresa, reconhecida, caminho_do_pdf), ...]
    """
    pasta_destino = Path(pasta_destino)
    if pasta_destino.exists():
        shutil.rmtree(pasta_destino, ignore_errors=True)
    pasta_destino.mkdir(parents=True, exist_ok=True)

    itens = []
    for indice, caminho in enumerate(caminhos_zip):
        caminho = Path(caminho)
        codigo, empresa, ok = empresa_do_zip(caminho.name)
        destino = pasta_destino / (str(indice) + "_" + re.sub(r"[^A-Za-z0-9]", "", caminho.stem)[:40])
        destino.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(caminho) as z:
            for nome in z.namelist():
                if nome.lower().endswith(".pdf") and not nome.endswith("/"):
                    z.extract(nome, destino)
        for pdf in sorted(destino.rglob("*.pdf")):
            itens.append((codigo, empresa, ok, pdf))
    return itens


def _executor(paralelo, trabalhadores):
    """Pool de processos quando possivel; senao roda sequencial."""
    if not paralelo:
        return None
    try:
        from concurrent.futures import ProcessPoolExecutor
        if sys.platform.startswith("win"):
            import multiprocessing
            executavel = sys.executable or ""
            if executavel.lower().endswith("pythonw.exe"):
                alternativo = executavel[:-len("pythonw.exe")] + "python.exe"
                if os.path.exists(alternativo):
                    multiprocessing.set_executable(alternativo)
        return ProcessPoolExecutor(max_workers=trabalhadores)
    except Exception:
        return None


def processar_lote(caminhos_zip, pasta_trabalho=None, progresso=None,
                   paralelo=True, trabalhadores=4):
    """
    Processa os ZIPs informados e devolve um ResultadoLote.
    progresso: funcao opcional progresso(feitos, total, texto)
    """
    def aviso(feitos, total, texto):
        if progresso:
            try:
                progresso(feitos, total, texto)
            except Exception:
                pass

    if pasta_trabalho is None:
        pasta_trabalho = Path(tempfile.gettempdir()) / "difal_lote_atual"

    resultado = ResultadoLote()
    resultado.zips = len(caminhos_zip)

    aviso(0, 1, "Extraindo os ZIPs...")
    itens = extrair_zips(caminhos_zip, pasta_trabalho)
    resultado.pdfs = len(itens)

    caminhos = [str(p) for (_, _, _, p) in itens]
    lidos = []
    total = len(caminhos)
    aviso(0, total, "Lendo os PDFs...")

    pool = _executor(paralelo, trabalhadores)
    if pool is None:
        for i, c in enumerate(caminhos, 1):
            lidos.append(processar_pdf(c))
            if i % 5 == 0 or i == total:
                aviso(i, total, "Lendo os PDFs...")
    else:
        try:
            with pool as executor:
                for i, dado in enumerate(executor.map(processar_pdf, caminhos, chunksize=4), 1):
                    lidos.append(dado)
                    if i % 5 == 0 or i == total:
                        aviso(i, total, "Lendo os PDFs...")
        except Exception:
            lidos = []
            for i, c in enumerate(caminhos, 1):
                lidos.append(processar_pdf(c))
                if i % 5 == 0 or i == total:
                    aviso(i, total, "Lendo os PDFs (modo seguro)...")

    vistos = {}
    for (codigo_emp, empresa, reconhecida, pdf), dado in zip(itens, lidos):
        registro = {
            "empresa": empresa,
            "codigo_empresa": codigo_emp,
            "arquivo": pdf.name,
            "tipo": dado["tipo"],
            "codigo": dado["codigo"],
            "valor": dado["valor"],
            "vencimento": dado["vencimento"],
            "valor_por_fallback": dado["valor_por_fallback"],
            "motivo": dado["motivo"],
        }

        if not dado["codigo"] or dado["valor"] is None:
            resultado.sem_codigo.append(registro)
            continue
        if not reconhecida:
            resultado.nao_autorizadas.append(registro)
            continue
        if dado["codigo"] in vistos:
            registro["duplicada_de"] = vistos[dado["codigo"]]
            resultado.duplicadas.append(registro)
            continue

        vistos[dado["codigo"]] = pdf.name
        resultado.guias.append(registro)

    resultado.guias.sort(key=lambda g: (g["empresa"], g["codigo"]))
    aviso(total, total, "Concluído")
    return resultado


# ------------------------------------------------------------ planilha

AZUL_CABECALHO = "1F3864"


def gerar_planilha(resultado, caminho_saida):
    """Gera o .xlsx com uma unica aba 'Guias GNRE' e 3 colunas."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Guias GNRE"

    cabecalhos = ["Empresa", "Código de Barras", "Valor"]
    for coluna, titulo in enumerate(cabecalhos, 1):
        celula = ws.cell(row=1, column=coluna, value=titulo)
        celula.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor=AZUL_CABECALHO)
        celula.alignment = Alignment(horizontal="center", vertical="center")

    for linha, guia in enumerate(resultado.guias, 2):
        c1 = ws.cell(row=linha, column=1, value=guia["empresa"])
        c2 = ws.cell(row=linha, column=2, value=str(guia["codigo"]))
        c2.data_type = "s"          # texto, preserva zeros a esquerda
        c2.number_format = "@"
        c3 = ws.cell(row=linha, column=3, value=round(guia["valor"], 2))
        c3.number_format = "#,##0.00"
        for c in (c1, c2, c3):
            c.font = Font(name="Arial", size=10)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 53
    ws.column_dimensions["C"].width = 13
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:" + get_column_letter(3) + str(max(1, len(resultado.guias) + 1))

    caminho_saida = Path(caminho_saida)
    caminho_saida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho_saida)
    return caminho_saida


# ------------------------------------------------------------ conferencia

def conferir(resultado):
    """Checagens obrigatorias antes da entrega. Devolve (ok, [mensagens])."""
    problemas = []

    for guia in resultado.guias:
        if len(guia["codigo"]) != 48 or not guia["codigo"].isdigit():
            problemas.append("Código fora do padrão de 48 dígitos: " + guia["arquivo"])
        elif not guia["codigo"].startswith("858"):
            problemas.append("Código não começa por 858: " + guia["arquivo"])
        if guia["valor"] is None or guia["valor"] <= 0:
            problemas.append("Valor inválido: " + guia["arquivo"])
        if guia["valor_por_fallback"]:
            problemas.append("Valor obtido por fallback (conferir no PDF): " + guia["arquivo"])

    codigos = [g["codigo"] for g in resultado.guias]
    if len(codigos) != len(set(codigos)):
        problemas.append("Há códigos repetidos na planilha final.")

    empresas_lote = set(g["codigo_empresa"] for g in resultado.guias)
    faltando = [EMPRESAS[c] for c in EMPRESAS if c not in empresas_lote]
    if faltando:
        problemas.append("Empresas sem guia neste lote: " + ", ".join(sorted(faltando)))

    return (len(problemas) == 0), problemas


def texto_relatorio(resultado):
    """Resumo em texto para a tela e para o log."""
    linhas = []
    linhas.append("RESUMO DO LOTE - " + resultado.inicio.strftime("%d/%m/%Y %H:%M"))
    linhas.append("")
    linhas.append("ZIPs processados ......... " + str(resultado.zips) + " de " + str(TOTAL_EMPRESAS))
    linhas.append("PDFs lidos ............... " + str(resultado.pdfs))
    linhas.append("Guias únicas exportadas .. " + str(len(resultado.guias)))
    linhas.append("Duplicatas removidas ..... " + str(len(resultado.duplicadas)))
    linhas.append("PDFs sem código .......... " + str(len(resultado.sem_codigo)))
    linhas.append("Fora da lista de empresas  " + str(len(resultado.nao_autorizadas)))
    tipos = resultado.por_tipo()
    linhas.append("Por tipo ................. GNRE " + str(tipos.get("GNRE", 0)) +
                  " | DARE-SP " + str(tipos.get("DARE-SP", 0)))
    linhas.append("TOTAL A PAGAR ............ R$ " +
                  ("{:,.2f}".format(resultado.total).replace(",", "X").replace(".", ",").replace("X", ".")))
    linhas.append("")
    linhas.append("POR EMPRESA")
    for empresa, dados in resultado.por_empresa().items():
        valor = "{:,.2f}".format(dados["total"]).replace(",", "X").replace(".", ",").replace("X", ".")
        linhas.append("  " + empresa.ljust(30) + str(dados["guias"]).rjust(4) + "   R$ " + valor.rjust(12))

    if resultado.duplicadas:
        linhas.append("")
        linhas.append("DUPLICATAS REMOVIDAS (mesmo código de barras)")
        for d in resultado.duplicadas:
            linhas.append("  " + d["empresa"] + " | " + d["arquivo"] + " = " + d.get("duplicada_de", ""))

    if resultado.sem_codigo:
        linhas.append("")
        linhas.append("PDFs SEM CÓDIGO (não entraram na planilha)")
        for s in resultado.sem_codigo:
            linhas.append("  " + s["empresa"] + " | " + s["arquivo"] + " | " + str(s["motivo"]))

    if resultado.nao_autorizadas:
        linhas.append("")
        linhas.append("EMPRESAS FORA DA LISTA (não entraram na planilha)")
        for n in resultado.nao_autorizadas:
            linhas.append("  " + n["empresa"] + " | " + n["arquivo"])

    ok, problemas = conferir(resultado)
    linhas.append("")
    linhas.append("CONFERÊNCIA: " + ("tudo certo" if ok else str(len(problemas)) + " ponto(s) de atenção"))
    for p in problemas:
        linhas.append("  - " + p)

    return "\n".join(linhas)


# ------------------------------------------------------------ banco C6

LIMITE_BANCO = 100          # o C6 aceita no maximo 100 pagamentos por arquivo
ABA_BANCO = "BOLETOS"
PRIMEIRA_LINHA_BANCO = 3
ULTIMA_LINHA_BANCO = 102


def _nome_seguro(texto):
    limpo = re.sub(r"[^A-Za-z0-9]+", "_", _sem_acento(texto)).strip("_")
    return limpo[:40] if limpo else "EMPRESA"


def _sem_acento(texto):
    t = unicodedata.normalize("NFD", str(texto))
    return "".join(c for c in t if unicodedata.category(c) != "Mn")


def _preencher_modelo(modelo, guias, data_pagamento, destino):
    """Copia o modelo do banco e preenche a aba BOLETOS sem mexer no resto."""
    from openpyxl import load_workbook

    if len(guias) > LIMITE_BANCO:
        raise ValueError("Um arquivo do banco não pode passar de " + str(LIMITE_BANCO) + " pagamentos.")

    wb = load_workbook(modelo)
    if ABA_BANCO not in wb.sheetnames:
        raise ValueError("O modelo do banco não tem a aba " + ABA_BANCO + ".")
    ws = wb[ABA_BANCO]

    # limpa as linhas de exemplo do modelo, preservando formatos
    for linha in range(PRIMEIRA_LINHA_BANCO, ULTIMA_LINHA_BANCO + 1):
        for coluna in range(1, 5):
            ws.cell(row=linha, column=coluna).value = None

    for indice, guia in enumerate(guias):
        linha = PRIMEIRA_LINHA_BANCO + indice
        celula = ws.cell(row=linha, column=1, value=str(guia["codigo"]))
        celula.data_type = "s"
        ws.cell(row=linha, column=2, value=round(guia["valor"], 2))
        ws.cell(row=linha, column=3, value=data_pagamento)
        ws.cell(row=linha, column=4, value=guia["empresa"])

    destino = Path(destino)
    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)
    return destino


def gerar_arquivos_banco(resultado, modelo, pasta_saida, data_pagamento,
                         por_empresa=True, limite=LIMITE_BANCO):
    """
    Gera os arquivos no layout do C6, com no maximo 'limite' pagamentos cada.
    por_empresa=True: nunca mistura empresas no mesmo arquivo.
    Devolve [(caminho, empresa, quantidade, total), ...]
    """
    modelo = Path(modelo)
    if not modelo.exists():
        raise FileNotFoundError("Modelo do banco não encontrado: " + str(modelo))

    pasta_saida = Path(pasta_saida)
    pasta_saida.mkdir(parents=True, exist_ok=True)

    if por_empresa:
        grupos = {}
        for guia in resultado.guias:
            grupos.setdefault(guia["empresa"], []).append(guia)
        grupos = sorted(grupos.items())
    else:
        grupos = [("TODAS", list(resultado.guias))]

    gerados = []
    for empresa, guias in grupos:
        blocos = [guias[i:i + limite] for i in range(0, len(guias), limite)]
        total_blocos = len(blocos)
        for numero, bloco in enumerate(blocos, 1):
            nome = ("C6_" + _nome_seguro(empresa) + "_" +
                    str(numero).zfill(2) + "de" + str(total_blocos).zfill(2) + ".xlsx")
            caminho = _preencher_modelo(modelo, bloco, data_pagamento, pasta_saida / nome)
            gerados.append((caminho, empresa, len(bloco), sum(g["valor"] for g in bloco)))
    return gerados


def texto_relatorio_banco(gerados, data_pagamento):
    linhas = ["ARQUIVOS PARA O BANCO C6 - pagamento em " + str(data_pagamento), ""]
    total_geral = 0.0
    total_guias = 0
    for caminho, empresa, quantidade, total in gerados:
        valor = "{:,.2f}".format(total).replace(",", "X").replace(".", ",").replace("X", ".")
        linhas.append("  " + Path(caminho).name.ljust(46) + str(quantidade).rjust(4) +
                      " guias   R$ " + valor.rjust(12))
        total_geral += total
        total_guias += quantidade
    valor = "{:,.2f}".format(total_geral).replace(",", "X").replace(".", ",").replace("X", ".")
    linhas.append("")
    linhas.append("  " + str(len(gerados)) + " arquivo(s)  ·  " + str(total_guias) +
                  " guias  ·  R$ " + valor)
    return "\n".join(linhas)


def salvar_json(resultado, caminho):
    """Guarda o lote bruto para auditoria posterior."""
    dados = {
        "data": resultado.inicio.isoformat(),
        "zips": resultado.zips,
        "pdfs": resultado.pdfs,
        "guias": resultado.guias,
        "duplicadas": resultado.duplicadas,
        "sem_codigo": resultado.sem_codigo,
        "nao_autorizadas": resultado.nao_autorizadas,
    }
    Path(caminho).write_text(json.dumps(dados, ensure_ascii=False, indent=2), encoding="utf-8")
